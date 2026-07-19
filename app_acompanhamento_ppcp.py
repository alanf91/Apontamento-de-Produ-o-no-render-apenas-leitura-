# VERSÃO 4.4: exibe o Calendário Semanal somente de segunda a sexta-feira.
from __future__ import annotations

import hmac
import re
import shutil
import subprocess
import sys
import unicodedata
from datetime import date, datetime, time, timedelta
from html import escape
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

NOME_ABA_PADRAO = "5_ACOMPANHAMENTO"
ARQUIVO_PADRAO = "PPCP_ACOMPAN.xlsx"
VERSAO_APP = "4.4"
LOGIN_APONTAMENTO = "pcp"
SENHA_APONTAMENTO = "pcp"



# -----------------------------------------------------------------------------
# Utilidades gerais
# -----------------------------------------------------------------------------
def normalizar(texto: Any) -> str:
    """Normaliza textos para comparar nomes de colunas sem acento/caixa."""
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^a-z0-9%]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def texto_limpo(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def numero_limpo(valor: Any) -> Any:
    if valor is None or valor == "":
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return int(valor)
    return valor


def valor_para_data(valor: Any) -> date | None:
    """Converte datas reais ou serial Excel para date."""
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, (int, float)) and 20000 <= valor <= 60000:
        try:
            return from_excel(valor).date()
        except Exception:
            return None
    try:
        convertido = pd.to_datetime(valor, dayfirst=True, errors="coerce")
        if pd.notna(convertido):
            return convertido.date()
    except Exception:
        pass
    return None


def data_br(valor: Any) -> str:
    d = valor_para_data(valor)
    return d.strftime("%d/%m/%Y") if d else ""


def to_num(serie: pd.Series) -> pd.Series:
    return pd.to_numeric(serie, errors="coerce").fillna(0)


def fmt_num(valor: Any, casas: int = 0) -> str:
    try:
        numero = float(valor)
    except Exception:
        numero = 0
    texto = f"{numero:,.{casas}f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def chave_ordenacao_texto(texto: Any) -> list[Any]:
    """Ordena textos com números de forma natural: 6 antes de 10."""
    partes = re.split(r"(\d+)", texto_limpo(texto))
    return [int(p) if p.isdigit() else p.lower() for p in partes]


def opcoes_unicas(df: pd.DataFrame, coluna: str) -> list[str]:
    """Retorna opções únicas limpas, sem depender dos filtros atuais."""
    if coluna not in df.columns:
        return []
    vistos: set[str] = set()
    opcoes: list[str] = []
    for valor in df[coluna].dropna().tolist():
        texto = texto_limpo(valor)
        if texto and texto not in vistos:
            vistos.add(texto)
            opcoes.append(texto)
    return sorted(opcoes, key=chave_ordenacao_texto)


# -----------------------------------------------------------------------------
# Leitura e gravação da planilha
# -----------------------------------------------------------------------------
def localizar_colunas(ws) -> dict[str, int]:
    """Localiza colunas pelo cabeçalho da linha 1."""
    cabecalhos = {normalizar(ws.cell(row=1, column=c).value): c for c in range(1, ws.max_column + 1)}

    candidatos = {
        "data_programada": ["data programada", "programada", "data planejamento", "data planejada"],
        # Data Entrega é usada especificamente para medir atraso contra a data de hoje.
        # Se a planilha não tiver esta coluna, o sistema usa Data Programada como fallback.
        "data_entrega": ["data entrega", "data de entrega", "entrega", "dt entrega", "prazo entrega", "prazo de entrega"],
        "lote": ["op lote", "lote", "op/lote"],
        "cliente": ["cliente pedido", "cliente/pedido"],
        "produto": ["produto equipamento", "produto/equipamento"],
        "codigo": ["codigo peca", "cod peca", "codigo"],
        "descricao": ["descricao peca", "descricao"],
        "operacao": ["operacao", "operação"],
        "setor": ["setor"],
        "maquina": ["maquina posto", "maquina/posto", "máquina/posto"],
        "qtde_programada": ["qtde programada", "quantidade programada"],
        "setup_padrao": ["setup padrao min", "setup padrão min", "setup padrao"],
        "tempo_unitario": ["tempo unit padrao min peca", "tempo unit padrão min peça", "tempo unit padrao", "tempo ciclo", "tempo de ciclo"],
        "tempo_programado": ["tempo programado min", "tempo programado"],
        "data_realizada": ["data realizada"],
        "qtde_realizada": ["qtde realizada", "quantidade realizada"],
        "inicio_real": ["inicio real", "início real"],
        "fim_real": ["fim real"],
        "tempo_realizado": ["tempo realizado min", "tempo realizado"],
        "refugo_retrabalho": ["refugo retrabalho", "refugo/retrabalho"],
        "percentual": ["% realizado", "percentual realizado"],
        "status": ["status"],
        "atraso": ["atraso?", "atraso"],
    }

    colunas: dict[str, int] = {}
    for chave, nomes in candidatos.items():
        for nome in nomes:
            nome_norm = normalizar(nome)
            if nome_norm in cabecalhos:
                colunas[chave] = cabecalhos[nome_norm]
                break

    # Compatibilidade: algumas bases antigas têm apenas Data Entrega ou apenas Data Programada.
    if "data_programada" not in colunas and "data_entrega" in colunas:
        colunas["data_programada"] = colunas["data_entrega"]
    if "data_entrega" not in colunas and "data_programada" in colunas:
        colunas["data_entrega"] = colunas["data_programada"]

    obrigatorias = [
        "data_programada",
        "lote",
        "setor",
        "qtde_programada",
        "data_realizada",
        "qtde_realizada",
    ]
    faltando = [c for c in obrigatorias if c not in colunas]
    if faltando:
        encontrados = ", ".join(str(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1))
        raise ValueError(
            "Não localizei estas colunas obrigatórias: "
            + ", ".join(faltando)
            + "\n\nCabeçalhos encontrados: "
            + encontrados
        )
    return colunas


def valor_celula(ws, linha: int, colunas: dict[str, int], chave: str) -> Any:
    col = colunas.get(chave)
    if not col:
        return ""
    return ws.cell(linha, col).value


def float_seguro(valor: Any) -> float:
    if valor in (None, ""):
        return 0.0
    try:
        return float(valor)
    except Exception:
        try:
            return float(str(valor).strip().replace(",", "."))
        except Exception:
            return 0.0


def carregar_tempos_cadastro(wb_valores) -> dict[str, tuple[float, float]]:
    """Lê os tempos técnicos da aba de cadastro quando as fórmulas do acompanhamento não têm cache."""
    nome_aba_cadastro = "01_Cad_Pecas"
    if nome_aba_cadastro not in wb_valores.sheetnames:
        return {}

    ws_cad = wb_valores[nome_aba_cadastro]
    cabecalhos = {
        normalizar(ws_cad.cell(row=1, column=c).value): c
        for c in range(1, ws_cad.max_column + 1)
    }
    col_codigo = cabecalhos.get(normalizar("Codigo Peca"))
    col_setup = cabecalhos.get(normalizar("Tempo Setup Padrao (min)"))
    col_unitario = cabecalhos.get(normalizar("Tempo Unit Padrao (min/peca)"))
    if not col_codigo or not col_setup or not col_unitario:
        return {}

    lookup: dict[str, tuple[float, float]] = {}
    for linha in range(2, ws_cad.max_row + 1):
        codigo = texto_limpo(ws_cad.cell(linha, col_codigo).value)
        if not codigo:
            continue
        setup = float_seguro(ws_cad.cell(linha, col_setup).value)
        unitario = float_seguro(ws_cad.cell(linha, col_unitario).value)
        # MATCH do Excel retorna a primeira ocorrência; setdefault preserva o mesmo comportamento.
        lookup.setdefault(codigo, (setup, unitario))
    return lookup


def carregar_dados(caminho: Path, nome_aba: str) -> tuple[pd.DataFrame, dict[str, int]]:
    # Abre duas leituras: uma preserva fórmulas/cabeçalhos e outra traz os valores calculados
    # que o Excel deixou gravados. Isso é importante para Tempo Programado (min), que é fórmula.
    wb = load_workbook(caminho, data_only=False)
    wb_valores = load_workbook(caminho, data_only=True)
    if nome_aba not in wb.sheetnames:
        wb.close()
        wb_valores.close()
        raise ValueError(f'A aba "{nome_aba}" não existe. Abas encontradas: {", ".join(wb.sheetnames)}')
    ws = wb[nome_aba]
    ws_valores = wb_valores[nome_aba]
    col = localizar_colunas(ws)
    tempos_cadastro = carregar_tempos_cadastro(wb_valores)

    linhas = []
    for linha_excel in range(2, ws.max_row + 1):
        lote = ws_valores.cell(linha_excel, col["lote"]).value
        setor = ws_valores.cell(linha_excel, col["setor"]).value
        qtd_prog = ws_valores.cell(linha_excel, col["qtde_programada"]).value
        if lote in (None, "") and setor in (None, "") and qtd_prog in (None, ""):
            continue

        data_prog_valor = ws_valores.cell(linha_excel, col["data_programada"]).value
        data_entrega_valor = ws_valores.cell(linha_excel, col["data_entrega"]).value if "data_entrega" in col else data_prog_valor
        data_real_valor = ws_valores.cell(linha_excel, col["data_realizada"]).value
        data_prog = valor_para_data(data_prog_valor)
        data_entrega = valor_para_data(data_entrega_valor)
        data_real = valor_para_data(data_real_valor)
        qtde_real = ws_valores.cell(linha_excel, col["qtde_realizada"]).value

        codigo_valor = valor_celula(ws_valores, linha_excel, col, "codigo")
        setup_valor = valor_celula(ws_valores, linha_excel, col, "setup_padrao")
        tempo_unitario_valor = valor_celula(ws_valores, linha_excel, col, "tempo_unitario")
        tempo_programado_valor = valor_celula(ws_valores, linha_excel, col, "tempo_programado")

        # O openpyxl não calcula fórmulas. Se a planilha veio sem cache calculado,
        # busca Setup e Tempo Unitário diretamente no cadastro técnico de peças.
        tempos_tecnicos = tempos_cadastro.get(texto_limpo(codigo_valor))
        if tempos_tecnicos:
            if setup_valor in (None, ""):
                setup_valor = tempos_tecnicos[0]
            if tempo_unitario_valor in (None, ""):
                tempo_unitario_valor = tempos_tecnicos[1]

        tempo_calculado = float_seguro(setup_valor) + (
            float_seguro(tempo_unitario_valor) * float_seguro(qtd_prog)
        )
        if float_seguro(tempo_programado_valor) <= 0 and tempo_calculado > 0:
            tempo_programado_valor = tempo_calculado

        linhas.append(
            {
                "Selecionar": False,
                "Linha Excel": linha_excel,
                "Data Programada": data_br(data_prog_valor),
                "Data Programada Valor": data_prog,
                "Data Entrega": data_br(data_entrega_valor),
                "Data Entrega Valor": data_entrega,
                "OP/Lote": texto_limpo(lote),
                "Cliente/Pedido": texto_limpo(valor_celula(ws_valores, linha_excel, col, "cliente")),
                "Produto/Equipamento": texto_limpo(valor_celula(ws_valores, linha_excel, col, "produto")),
                "Código Peça": texto_limpo(codigo_valor),
                "Descrição Peça": texto_limpo(valor_celula(ws_valores, linha_excel, col, "descricao")),
                "Operação": texto_limpo(valor_celula(ws_valores, linha_excel, col, "operacao")),
                "Setor": texto_limpo(setor),
                "Máquina/Posto": texto_limpo(valor_celula(ws_valores, linha_excel, col, "maquina")),
                "Qtde Programada": numero_limpo(qtd_prog),
                "Setup Padrão (min)": numero_limpo(setup_valor),
                "Tempo Unit Padrão (min/peça)": numero_limpo(tempo_unitario_valor),
                "Tempo Programado (min)": numero_limpo(tempo_programado_valor),
                "Data Realizada": data_br(data_real_valor),
                "Data Realizada Valor": data_real,
                "Qtde Realizada": numero_limpo(qtde_real),
                "Início Real": texto_limpo(valor_celula(ws_valores, linha_excel, col, "inicio_real")),
                "Fim Real": texto_limpo(valor_celula(ws_valores, linha_excel, col, "fim_real")),
                "Tempo Realizado (min)": numero_limpo(valor_celula(ws_valores, linha_excel, col, "tempo_realizado")),
                "Refugo/Retrabalho": numero_limpo(valor_celula(ws_valores, linha_excel, col, "refugo_retrabalho")),
                "Situação": "Realizado" if data_real is not None or qtde_real not in (None, "") else "Pendente",
            }
        )
    wb.close()
    wb_valores.close()
    return pd.DataFrame(linhas), col


@st.cache_data(show_spinner=False)
def carregar_dados_cacheado(
    caminho_texto: str,
    nome_aba: str,
    modificacao_ns: int,
    tamanho_arquivo: int,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """
    Mantém a planilha em memória entre as interações do Streamlit.

    Os parâmetros de modificação e tamanho fazem parte da chave do cache.
    Quando o Excel é salvo ou substituído, o cache é renovado automaticamente.
    """
    del modificacao_ns, tamanho_arquivo
    return carregar_dados(Path(caminho_texto), nome_aba)


def obter_dados_cacheados(caminho: Path, nome_aba: str) -> tuple[pd.DataFrame, dict[str, int]]:
    estatistica = caminho.stat()
    return carregar_dados_cacheado(
        str(caminho.resolve()),
        nome_aba,
        estatistica.st_mtime_ns,
        estatistica.st_size,
    )


def salvar_realizacao(
    caminho: Path,
    nome_aba: str,
    linhas_excel: list[int],
    data_realizada: date,
    preencher_qtde: bool,
    atualizar_status: bool,
    criar_backup: bool,
) -> Path | None:
    if criar_backup:
        backup = caminho.with_name(f"{caminho.stem}_backup_{datetime.now():%Y%m%d_%H%M%S}{caminho.suffix}")
        shutil.copy2(caminho, backup)
    else:
        backup = None

    wb = load_workbook(caminho, data_only=False)
    ws = wb[nome_aba]
    col = localizar_colunas(ws)

    data_excel = datetime.combine(data_realizada, time.min)
    letra_qtd_prog = get_column_letter(col["qtde_programada"])
    letra_qtd_real = get_column_letter(col["qtde_realizada"])

    for linha in linhas_excel:
        cel_data = ws.cell(linha, col["data_realizada"])
        cel_data.value = data_excel
        cel_data.number_format = "dd/mm/yyyy"

        if preencher_qtde:
            qtde_programada = ws.cell(linha, col["qtde_programada"]).value
            cel_qtd = ws.cell(linha, col["qtde_realizada"])
            cel_qtd.value = qtde_programada
            cel_qtd.number_format = "0"

        if "percentual" in col:
            ws.cell(linha, col["percentual"]).value = (
                f'=IF(OR(${letra_qtd_prog}{linha}="",${letra_qtd_real}{linha}=""),"",${letra_qtd_real}{linha}/${letra_qtd_prog}{linha})'
            )
            ws.cell(linha, col["percentual"]).number_format = "0%"

        if atualizar_status and "status" in col:
            ws.cell(linha, col["status"]).value = "Concluído"

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(caminho)
    wb.close()
    return backup


# -----------------------------------------------------------------------------
# Tela de apontamento
# -----------------------------------------------------------------------------
def filtrar_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    st.sidebar.header("Filtros do apontamento")

    # As opções de setor agora vêm da base completa, e não apenas do resultado
    # já filtrado. Isso evita sumir setores como "6 PINTURA UV" quando outros
    # filtros estiverem ativos ou quando o setor aparecer mais abaixo na planilha.
    todos_setores = opcoes_unicas(df, "Setor")
    todas_maquinas = opcoes_unicas(df, "Máquina/Posto")

    with st.sidebar.expander("Ver setores encontrados"):
        st.caption(f"{len(todos_setores)} setores encontrados na planilha carregada.")
        if todos_setores:
            st.write(", ".join(todos_setores))
        else:
            st.write("Nenhum setor encontrado.")

    situacao = st.sidebar.radio("Situação", ["Pendentes", "Realizados", "Todos"], horizontal=True)
    filtrado = df.copy()
    if situacao == "Pendentes":
        filtrado = filtrado[filtrado["Situação"] == "Pendente"]
    elif situacao == "Realizados":
        filtrado = filtrado[filtrado["Situação"] == "Realizado"]

    busca_setor = st.sidebar.text_input("Buscar setor digitando", placeholder="Ex.: PINTURA, UV, 6")
    setores_para_exibir = todos_setores
    if busca_setor:
        busca_norm = normalizar(busca_setor)
        setores_para_exibir = [s for s in todos_setores if busca_norm in normalizar(s)]
        filtrado = filtrado[filtrado["Setor"].astype(str).apply(lambda x: busca_norm in normalizar(x))]

    setor_sel = st.sidebar.multiselect(
        "Filtrar por setor",
        setores_para_exibir,
        placeholder="Selecione um ou mais setores",
    )
    if setor_sel:
        filtrado = filtrado[filtrado["Setor"].isin(setor_sel)]

    busca_lote = st.sidebar.text_input("Buscar lote", placeholder="Ex.: 10176")
    if busca_lote:
        filtrado = filtrado[filtrado["OP/Lote"].str.contains(busca_lote, case=False, na=False)]

    # A lista de máquinas também vem da base completa para não desaparecer ao trocar filtros.
    maquinas_para_exibir = todas_maquinas
    if setor_sel:
        maquinas_para_exibir = opcoes_unicas(df[df["Setor"].isin(setor_sel)], "Máquina/Posto")
    elif busca_setor:
        busca_norm = normalizar(busca_setor)
        maquinas_para_exibir = opcoes_unicas(df[df["Setor"].astype(str).apply(lambda x: busca_norm in normalizar(x))], "Máquina/Posto")

    maquinas_sel = st.sidebar.multiselect("Máquina/Posto", maquinas_para_exibir)
    if maquinas_sel:
        filtrado = filtrado[filtrado["Máquina/Posto"].isin(maquinas_sel)]

    texto = st.sidebar.text_input("Buscar peça/produto/cliente", placeholder="Digite parte do texto")
    if texto:
        campos = ["Cliente/Pedido", "Produto/Equipamento", "Código Peça", "Descrição Peça", "Operação"]
        mascara = pd.Series(False, index=filtrado.index)
        for campo in campos:
            mascara = mascara | filtrado[campo].astype(str).str.contains(texto, case=False, na=False)
        filtrado = filtrado[mascara]

    return filtrado


def renderizar_apontamento(df: pd.DataFrame, caminho: Path, nome_aba: str) -> None:
    st.title("Acompanhamento de Produção - PPCP")
    st.caption(f"Sistema versão {VERSAO_APP}")
    st.caption(
        "Modo rápido: marque quantas linhas precisar e clique em Salvar. "
        "A página não é recarregada a cada caixa selecionada."
    )

    total = len(df)
    pendentes = int((df["Situação"] == "Pendente").sum())
    realizados = int((df["Situação"] == "Realizado").sum())

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total de linhas", total)
    c2.metric("Pendentes", pendentes)
    c3.metric("Realizados", realizados)
    c4.metric("Arquivo", caminho.name)

    filtrado = filtrar_dataframe(df).copy()

    if filtrado.empty:
        st.warning("Nenhum item encontrado para os filtros selecionados.")
        return

    # Paginação reduz a quantidade de células que o navegador precisa renderizar.
    col_pagina1, col_pagina2 = st.columns([1, 1])
    opcoes_por_pagina = [100, 200, 300, 500, 1000]
    itens_por_pagina = col_pagina1.selectbox(
        "Itens por página",
        opcoes_por_pagina,
        index=1,
        help="Quanto menor este valor, mais rápida fica a seleção das linhas.",
        key="apontamento_itens_por_pagina",
    )

    total_paginas = max(1, (len(filtrado) + itens_por_pagina - 1) // itens_por_pagina)
    pagina_guardada = int(st.session_state.get("apontamento_pagina", 1))
    if pagina_guardada < 1 or pagina_guardada > total_paginas:
        st.session_state["apontamento_pagina"] = 1
    pagina_atual = col_pagina2.number_input(
        "Página",
        min_value=1,
        max_value=total_paginas,
        step=1,
        key="apontamento_pagina",
    )

    inicio = (int(pagina_atual) - 1) * itens_por_pagina
    fim = min(inicio + itens_por_pagina, len(filtrado))
    pagina_df = filtrado.iloc[inicio:fim].copy().reset_index(drop=True)

    st.info(
        f"Exibindo itens **{inicio + 1} a {fim}** de **{len(filtrado)}** filtrados "
        f"— página **{int(pagina_atual)} de {total_paginas}**."
    )

    mapa_linhas_excel_pagina = pagina_df["Linha Excel"].astype(int).reset_index(drop=True)
    todas_linhas_filtradas = filtrado["Linha Excel"].astype(int).tolist()

    colunas_visiveis = [
        "Selecionar",
        "Data Programada",
        "Data Entrega",
        "OP/Lote",
        "Setor",
        "Máquina/Posto",
        "Produto/Equipamento",
        "Código Peça",
        "Descrição Peça",
        "Qtde Programada",
        "Data Realizada",
        "Qtde Realizada",
        "Situação",
    ]
    tabela = pagina_df[[c for c in colunas_visiveis if c in pagina_df.columns]].copy()
    tabela["Selecionar"] = False

    # Dentro do formulário, as marcações ficam no navegador e só são enviadas
    # quando o usuário clica em Salvar. Isso elimina uma execução completa por clique.
    with st.form("form_apontamento_rapido", clear_on_submit=False):
        st.subheader("Lançamento rápido")
        col_data, col_qtd, col_status, col_backup = st.columns([1, 1.2, 1.2, 1])
        data_realizada = col_data.date_input(
            "Data realizada",
            value=date.today(),
            format="DD/MM/YYYY",
        )
        preencher_qtde = col_qtd.checkbox(
            "Qtde realizada = Qtde programada",
            value=True,
        )
        atualizar_status = col_status.checkbox(
            "Marcar status como Concluído",
            value=True,
        )
        criar_backup = col_backup.checkbox(
            "Criar backup antes de salvar",
            value=True,
        )

        salvar_todos_filtrados = st.checkbox(
            "Salvar todos os itens filtrados, incluindo as outras páginas",
            value=False,
            help=(
                "Quando marcado, o sistema ignora as caixas individuais e salva "
                "todos os itens encontrados pelos filtros atuais."
            ),
        )

        st.caption(
            "Marque as caixas livremente. O botão fica disponível o tempo todo; "
            "as seleções serão conferidas somente ao clicar em Salvar."
        )

        editada = st.data_editor(
            tabela,
            use_container_width=True,
            hide_index=True,
            height=520,
            disabled=[c for c in tabela.columns if c != "Selecionar"],
            column_config={
                "Selecionar": st.column_config.CheckboxColumn("Selecionar"),
                "Descrição Peça": st.column_config.TextColumn("Descrição Peça", width="large"),
                "Produto/Equipamento": st.column_config.TextColumn("Produto/Equipamento", width="medium"),
            },
        )

        salvar = st.form_submit_button(
            "Salvar apontamentos selecionados",
            type="primary",
            use_container_width=True,
        )

    if salvar:
        if salvar_todos_filtrados:
            linhas_selecionadas = todas_linhas_filtradas
        else:
            indices_selecionados = editada.index[editada["Selecionar"]].tolist()
            linhas_selecionadas = (
                mapa_linhas_excel_pagina.loc[indices_selecionados].astype(int).tolist()
                if indices_selecionados
                else []
            )

        if not linhas_selecionadas:
            st.warning("Nenhuma linha foi selecionada. Marque pelo menos um item e tente novamente.")
        else:
            try:
                with st.spinner(f"Salvando {len(linhas_selecionadas)} apontamento(s)..."):
                    backup = salvar_realizacao(
                        caminho=caminho,
                        nome_aba=nome_aba,
                        linhas_excel=linhas_selecionadas,
                        data_realizada=data_realizada,
                        preencher_qtde=preencher_qtde,
                        atualizar_status=atualizar_status,
                        criar_backup=criar_backup,
                    )
                # Força a próxima leitura a refletir o Excel recém-salvo.
                carregar_dados_cacheado.clear()
                st.success(f"Salvo com sucesso em {len(linhas_selecionadas)} linha(s).")
                if backup:
                    st.info(f"Backup criado: {backup.name}")
                st.rerun()
            except PermissionError:
                st.error("Não consegui salvar. Feche a planilha no Excel e tente novamente.")
            except Exception as exc:
                st.error(f"Erro ao salvar: {exc}")

    st.divider()
    with open(caminho, "rb") as f:
        st.download_button(
            "Baixar planilha atualizada",
            data=f.read(),
            file_name=caminho.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with st.expander("Como usar o modo rápido"):
        st.markdown(
            """
1. Use os filtros de setor, lote, máquina ou texto.
2. Escolha quantos itens deseja exibir por página.
3. Marque as linhas necessárias sem esperar recarregamentos.
4. Para salvar todo o resultado do filtro, marque **Salvar todos os itens filtrados**.
5. Clique em **Salvar apontamentos selecionados**.

O formulário envia todas as marcações de uma vez. Por isso, a quantidade selecionada não é recalculada a cada clique.
            """
        )

# -----------------------------------------------------------------------------
# Tela de produção por setor
# -----------------------------------------------------------------------------
def preparar_base_producao_setor(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara os campos usados na comparação por setor.

    Regra do realizado:
    - usa a Qtde Realizada quando ela foi informada;
    - para apontamentos antigos com Data Realizada preenchida e quantidade vazia,
      considera a Qtde Programada como realizada.
    """
    base = df.copy()
    base["Setor"] = base["Setor"].fillna("").astype(str).str.strip().replace("", "Sem setor")
    base["Qtde Programada Num"] = to_num(base["Qtde Programada"]).clip(lower=0)

    qtde_real_original = base["Qtde Realizada"].copy()
    base["Qtde Realizada Num"] = to_num(qtde_real_original).clip(lower=0)
    base["Data Programada DT"] = pd.to_datetime(base["Data Programada Valor"], errors="coerce")
    base["Data Realizada DT"] = pd.to_datetime(base["Data Realizada Valor"], errors="coerce")

    texto_qtde_real = qtde_real_original.fillna("").astype(str).str.strip()
    base["Qtde Realizada Vazia?"] = texto_qtde_real.eq("")
    base["Realizado por fallback?"] = (
        base["Data Realizada DT"].notna()
        & base["Qtde Realizada Vazia?"]
        & (base["Qtde Programada Num"] > 0)
    )
    base["Qtde Realizada Considerada Num"] = base["Qtde Realizada Num"]
    base.loc[
        base["Realizado por fallback?"],
        "Qtde Realizada Considerada Num",
    ] = base.loc[
        base["Realizado por fallback?"],
        "Qtde Programada Num",
    ]
    return base


def intervalo_datas_disponiveis(base: pd.DataFrame) -> tuple[date, date] | None:
    datas = pd.concat(
        [base["Data Programada DT"], base["Data Realizada DT"]],
        ignore_index=True,
    ).dropna()
    if datas.empty:
        return None
    return datas.min().date(), datas.max().date()


def escolher_filtros_producao_setor(
    base: pd.DataFrame,
) -> tuple[list[str], date | None, date | None, str]:
    st.subheader("Filtros")
    setores = opcoes_unicas(base, "Setor")

    col_setor, col_visao = st.columns([1.4, 1])
    setores_sel = col_setor.multiselect(
        "Setor",
        setores,
        placeholder="Vazio = todos os setores",
        key="producao_setores",
    )
    modo = col_visao.radio(
        "Visualização",
        ["Dia específico", "Período personalizado", "Período completo"],
        horizontal=False,
        key="producao_modo_data",
    )

    intervalo = intervalo_datas_disponiveis(base)
    if intervalo is None:
        st.warning("Não existem datas programadas ou realizadas válidas na base.")
        return setores_sel, None, None, modo

    data_min, data_max = intervalo
    inicio: date
    fim: date

    if modo == "Dia específico":
        hoje = date.today()
        data_padrao = hoje if data_min <= hoje <= data_max else data_max
        dia = st.date_input(
            "Dia analisado",
            value=data_padrao,
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="producao_dia",
        )
        inicio = dia
        fim = dia
    elif modo == "Período personalizado":
        periodo = st.date_input(
            "Período analisado",
            value=(data_min, data_max),
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="producao_periodo",
        )
        if isinstance(periodo, (tuple, list)) and len(periodo) == 2:
            inicio, fim = periodo
        elif isinstance(periodo, (tuple, list)) and len(periodo) == 1:
            inicio = fim = periodo[0]
        else:
            inicio = fim = periodo
    else:
        inicio, fim = data_min, data_max
        st.caption(f"Período completo da base: {inicio:%d/%m/%Y} a {fim:%d/%m/%Y}")

    return setores_sel, inicio, fim, modo


def entre_datas(serie: pd.Series, inicio: date, fim: date) -> pd.Series:
    return serie.notna() & (serie.dt.date >= inicio) & (serie.dt.date <= fim)


def montar_resumo_producao_setor(
    base: pd.DataFrame,
    setores_sel: list[str],
    inicio: date,
    fim: date,
) -> pd.DataFrame:
    """
    Compara o programado e o realizado das MESMAS linhas programadas no período.

    Exemplo: ao selecionar 17/07/2026, primeiro são localizadas todas as linhas
    cuja Data Programada é 17/07/2026. Depois:
    - Programado = soma da Qtde Programada dessas linhas;
    - Realizado = soma da Qtde Realizada apontada nessas mesmas linhas.
    """
    setores_base = setores_sel if setores_sel else opcoes_unicas(base, "Setor")
    base_setores = base[base["Setor"].isin(setores_base)].copy()
    base_periodo = base_setores[
        entre_datas(base_setores["Data Programada DT"], inicio, fim)
    ].copy()

    programado = (
        base_periodo.groupby("Setor")["Qtde Programada Num"]
        .sum()
        .rename("Programado")
    )
    realizado = (
        base_periodo.groupby("Setor")["Qtde Realizada Considerada Num"]
        .sum()
        .rename("Realizado")
    )

    resumo = pd.DataFrame({"Setor": setores_base}).set_index("Setor")
    resumo = resumo.join(programado).join(realizado).fillna(0).reset_index()
    resumo["Saldo"] = resumo["Programado"] - resumo["Realizado"]
    resumo["% Realizado"] = 0.0
    mask_programado = resumo["Programado"] > 0
    resumo.loc[mask_programado, "% Realizado"] = (
        resumo.loc[mask_programado, "Realizado"]
        / resumo.loc[mask_programado, "Programado"]
        * 100
    )

    def situacao(linha: pd.Series) -> str:
        programado_val = float(linha["Programado"])
        realizado_val = float(linha["Realizado"])
        percentual = float(linha["% Realizado"])
        if programado_val <= 0 and realizado_val <= 0:
            return "Sem programação"
        if realizado_val >= programado_val and programado_val > 0:
            return "Programado atingido"
        if percentual >= 90:
            return "Próximo do programado"
        return "Abaixo do programado"

    resumo["Situação"] = resumo.apply(situacao, axis=1)
    return resumo.sort_values(["Programado", "Realizado", "Setor"], ascending=[False, False, True])


def montar_producao_diaria(
    base: pd.DataFrame,
    setores_sel: list[str],
    inicio: date,
    fim: date,
) -> pd.DataFrame:
    """
    Monta o resultado diário pela Data Programada.

    Em cada dia, o realizado representa quanto foi apontado nas linhas que estavam
    programadas para aquele mesmo dia, mesmo que a Data Realizada seja diferente.
    """
    setores_base = setores_sel if setores_sel else opcoes_unicas(base, "Setor")
    base_setores = base[base["Setor"].isin(setores_base)].copy()
    periodo = base_setores[
        entre_datas(base_setores["Data Programada DT"], inicio, fim)
    ].copy()

    programado_dia = (
        periodo.groupby(periodo["Data Programada DT"].dt.date)["Qtde Programada Num"]
        .sum()
        .rename("Programado")
    )
    realizado_dia = (
        periodo.groupby(periodo["Data Programada DT"].dt.date)["Qtde Realizada Considerada Num"]
        .sum()
        .rename("Realizado")
    )
    programado_dia.index = pd.to_datetime(programado_dia.index)
    realizado_dia.index = pd.to_datetime(realizado_dia.index)

    calendario = pd.date_range(inicio, fim, freq="D")
    diario = pd.DataFrame(index=calendario)
    diario.index.name = "Data"
    diario = diario.join(programado_dia.rename_axis("Data"), how="left")
    diario = diario.join(realizado_dia.rename_axis("Data"), how="left")
    diario = diario.fillna(0).reset_index()
    diario["Saldo"] = diario["Programado"] - diario["Realizado"]
    diario["% Realizado"] = 0.0
    mask = diario["Programado"] > 0
    diario.loc[mask, "% Realizado"] = (
        diario.loc[mask, "Realizado"] / diario.loc[mask, "Programado"] * 100
    )
    return diario


def montar_detalhamento_periodo(
    base: pd.DataFrame,
    setores_sel: list[str],
    inicio: date,
    fim: date,
    tipo: str,
) -> pd.DataFrame:
    """
    Detalha as linhas programadas no período.

    As duas abas usam a Data Programada para garantir que o programado e o realizado
    sejam auditáveis sobre exatamente o mesmo conjunto de linhas.
    """
    setores_base = setores_sel if setores_sel else opcoes_unicas(base, "Setor")
    dados = base[base["Setor"].isin(setores_base)].copy()
    dados = dados[entre_datas(dados["Data Programada DT"], inicio, fim)]
    dados = dados.sort_values(["Data Programada DT", "Setor", "OP/Lote"])

    if tipo == "programado":
        return dados[
            [
                "Data Programada",
                "Setor",
                "OP/Lote",
                "Produto/Equipamento",
                "Código Peça",
                "Descrição Peça",
                "Qtde Programada",
            ]
        ]

    detalhe = dados[
        [
            "Data Programada",
            "Data Realizada",
            "Setor",
            "OP/Lote",
            "Produto/Equipamento",
            "Código Peça",
            "Descrição Peça",
            "Qtde Programada",
            "Qtde Realizada",
            "Qtde Realizada Considerada Num",
            "Realizado por fallback?",
        ]
    ].copy()
    detalhe = detalhe.rename(
        columns={
            "Qtde Realizada Considerada Num": "Qtde Realizada Considerada",
            "Realizado por fallback?": "Quantidade assumida pela data?",
        }
    )
    return detalhe


def renderizar_producao_setor(df: pd.DataFrame, caminho: Path) -> None:
    st.title("Produção por Setor")
    st.caption(
        "O filtro seleciona as linhas pela Data Programada. "
        "O realizado mostra quanto foi apontado nessas mesmas linhas programadas."
    )

    base = preparar_base_producao_setor(df)
    setores_sel, inicio, fim, modo = escolher_filtros_producao_setor(base)
    if inicio is None or fim is None:
        return

    resumo = montar_resumo_producao_setor(base, setores_sel, inicio, fim)
    diario = montar_producao_diaria(base, setores_sel, inicio, fim)

    setores_base = setores_sel if setores_sel else opcoes_unicas(base, "Setor")
    base_periodo = base[
        base["Setor"].isin(setores_base)
        & entre_datas(base["Data Programada DT"], inicio, fim)
    ]
    qtd_fallback = int(base_periodo["Realizado por fallback?"].sum())
    if qtd_fallback > 0:
        st.info(
            f"{qtd_fallback} linha(s) possuem Data Realizada, mas a Qtde Realizada está vazia. "
            "Nessas linhas, o sistema considerou a Qtde Programada como realizada."
        )

    total_programado = float(resumo["Programado"].sum())
    total_realizado = float(resumo["Realizado"].sum())
    saldo = total_programado - total_realizado
    percentual = (total_realizado / total_programado * 100) if total_programado > 0 else 0.0
    setores_movimento = int(((resumo["Programado"] > 0) | (resumo["Realizado"] > 0)).sum())

    st.divider()
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Programado do período", fmt_num(total_programado, 2))
    k2.metric("Realizado do programado", fmt_num(total_realizado, 2))
    k3.metric("Saldo do programado", fmt_num(saldo, 2))
    k4.metric("% atendimento", f"{percentual:.1f}%".replace(".", ","))
    k5.metric("Setores com movimento", fmt_num(setores_movimento))

    if modo == "Dia específico":
        st.caption(
            f"Resultado das linhas programadas para {inicio:%d/%m/%Y}. "
            "O realizado é o apontamento dessas mesmas linhas."
        )
    else:
        st.caption(f"Resultado acumulado de {inicio:%d/%m/%Y} a {fim:%d/%m/%Y}.")

    aba_setor, aba_dia, aba_detalhes = st.tabs(
        ["Comparação por setor", "Resultado por dia", "Detalhamento"]
    )

    with aba_setor:
        st.subheader("Programado x realizado por setor")
        somente_movimento = st.checkbox(
            "Mostrar somente setores com movimento no período",
            value=False,
            key="producao_somente_movimento",
        )
        exibicao = resumo.copy()
        if somente_movimento:
            exibicao = exibicao[(exibicao["Programado"] > 0) | (exibicao["Realizado"] > 0)]

        if exibicao.empty:
            st.info("Nenhum setor possui programação ou apontamento para os filtros selecionados.")
        else:
            st.bar_chart(exibicao.set_index("Setor")[["Programado", "Realizado"]])
            st.dataframe(
                exibicao,
                use_container_width=True,
                hide_index=True,
                height=520,
                column_config={
                    "Programado": st.column_config.NumberColumn("Programado", format="%.2f"),
                    "Realizado": st.column_config.NumberColumn("Realizado", format="%.2f"),
                    "Saldo": st.column_config.NumberColumn("Saldo", format="%.2f"),
                    "% Realizado": st.column_config.ProgressColumn(
                        "% Realizado",
                        min_value=0,
                        max_value=max(100.0, float(exibicao["% Realizado"].max())),
                        format="%.1f%%",
                    ),
                },
            )

            csv = exibicao.to_csv(index=False, sep=";", decimal=",", encoding="utf-8-sig")
            st.download_button(
                "Baixar resultado por setor em CSV",
                data=csv,
                file_name=f"producao_por_setor_{inicio:%Y%m%d}_{fim:%Y%m%d}.csv",
                mime="text/csv",
            )

    with aba_dia:
        st.subheader("Atendimento do programado por dia")
        diario_movimento = diario[(diario["Programado"] > 0) | (diario["Realizado"] > 0)].copy()
        if diario_movimento.empty:
            st.info("Não há programação no período selecionado.")
        else:
            grafico_diario = diario_movimento.copy()
            grafico_diario["Data"] = pd.to_datetime(grafico_diario["Data"])
            st.line_chart(grafico_diario.set_index("Data")[["Programado", "Realizado"]])

            tabela_diaria = diario_movimento.copy()
            tabela_diaria["Data"] = pd.to_datetime(tabela_diaria["Data"]).dt.strftime("%d/%m/%Y")
            st.dataframe(
                tabela_diaria,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Programado": st.column_config.NumberColumn("Programado", format="%.2f"),
                    "Realizado": st.column_config.NumberColumn("Realizado", format="%.2f"),
                    "Saldo": st.column_config.NumberColumn("Saldo", format="%.2f"),
                    "% Realizado": st.column_config.ProgressColumn(
                        "% Realizado",
                        min_value=0,
                        max_value=max(100.0, float(tabela_diaria["% Realizado"].max())),
                        format="%.1f%%",
                    ),
                },
            )

    with aba_detalhes:
        st.subheader("Linhas que formam os totais")
        detalhe_prog = montar_detalhamento_periodo(base, setores_sel, inicio, fim, "programado")
        detalhe_real = montar_detalhamento_periodo(base, setores_sel, inicio, fim, "realizado")

        sub_prog, sub_real = st.tabs(["Programado", "Realizado"])
        with sub_prog:
            st.caption(f"{len(detalhe_prog)} linha(s) consideradas no programado.")
            st.dataframe(detalhe_prog, use_container_width=True, hide_index=True, height=460)
        with sub_real:
            st.caption(f"{len(detalhe_real)} linha(s) programadas no período, com o respectivo apontamento.")
            st.dataframe(detalhe_real, use_container_width=True, hide_index=True, height=460)

    st.divider()
    st.caption(f"Fonte dos dados: {caminho.name}")


# -----------------------------------------------------------------------------
# Tela pública: Calendário Semanal dos Setores
# -----------------------------------------------------------------------------
DIAS_SEMANA_PT = [
    "Segunda",
    "Terça",
    "Quarta",
    "Quinta",
    "Sexta",
]


def fmt_qtd(valor: Any) -> str:
    """Formata quantidade sem casas desnecessárias e com padrão brasileiro."""
    try:
        numero = float(valor)
    except Exception:
        numero = 0.0
    casas = 0 if numero.is_integer() else 2
    return fmt_num(numero, casas)


def preparar_base_calendario_semanal(df: pd.DataFrame) -> pd.DataFrame:
    """
    Classifica cada linha programada em três grupos mutuamente exclusivos.

    - Realizada no prazo: Data Realizada menor ou igual à Data Programada;
    - Realizada após: Data Realizada maior que a Data Programada;
    - Pendente: parte programada que ainda não possui conclusão datada.

    A quantidade classificada é limitada à quantidade programada da linha. Assim,
    no prazo + após + pendente sempre fecha com o programado e um apontamento acima
    do programado não distorce o percentual de atraso.
    """
    base = preparar_base_producao_setor(df)

    programada = base["Qtde Programada Num"].clip(lower=0)
    realizada = base["Qtde Realizada Considerada Num"].clip(lower=0)
    base["Qtde Finalizada Classificada Num"] = pd.concat(
        [programada, realizada], axis=1
    ).min(axis=1)

    # Sem Data Realizada não é possível provar se a conclusão ocorreu no prazo.
    # Por isso, a quantidade permanece pendente até a data ser apontada.
    tem_data_realizada = base["Data Realizada DT"].notna()
    tem_data_programada = base["Data Programada DT"].notna()
    finalizada_com_data = base["Qtde Finalizada Classificada Num"].where(
        tem_data_realizada & tem_data_programada,
        0.0,
    )

    no_prazo = (
        tem_data_realizada
        & tem_data_programada
        & (base["Data Realizada DT"] <= base["Data Programada DT"])
    )
    apos_prazo = (
        tem_data_realizada
        & tem_data_programada
        & (base["Data Realizada DT"] > base["Data Programada DT"])
    )

    base["Realizada no Prazo Num"] = finalizada_com_data.where(no_prazo, 0.0)
    base["Realizada Após Num"] = finalizada_com_data.where(apos_prazo, 0.0)
    base["Pendente Num"] = (
        base["Qtde Programada Num"]
        - base["Realizada no Prazo Num"]
        - base["Realizada Após Num"]
    ).clip(lower=0)

    return base


def limites_datas_programadas(base: pd.DataFrame) -> tuple[date, date] | None:
    datas = base["Data Programada DT"].dropna()
    if datas.empty:
        return None
    return datas.min().date(), datas.max().date()


def inicio_da_semana(data_referencia: date) -> date:
    """Retorna a segunda-feira da semana que contém a data informada."""
    return data_referencia - timedelta(days=data_referencia.weekday())


def montar_dados_calendario_semanal(
    base: pd.DataFrame,
    setores_sel: list[str],
    inicio: date,
    fim: date,
    somente_movimento: bool,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """Monta o detalhamento diário e o resumo semanal por setor."""
    todos_setores = opcoes_unicas(base, "Setor")
    setores_base = setores_sel if setores_sel else todos_setores

    periodo = base[
        base["Setor"].isin(setores_base)
        & entre_datas(base["Data Programada DT"], inicio, fim)
    ].copy()
    periodo["Data Programada Dia"] = periodo["Data Programada DT"].dt.date

    colunas_soma = [
        "Qtde Programada Num",
        "Realizada no Prazo Num",
        "Realizada Após Num",
        "Pendente Num",
    ]

    if periodo.empty:
        diario = pd.DataFrame(
            columns=["Setor", "Data Programada Dia", *colunas_soma, "% Atraso"]
        )
        resumo = pd.DataFrame(
            {
                "Setor": setores_base,
                "Programada": 0.0,
                "Realizada no Prazo": 0.0,
                "Realizada Após": 0.0,
                "Pendente": 0.0,
                "% Atraso Semanal": 0.0,
            }
        )
        setores_exibicao = [] if somente_movimento else setores_base
        return diario, resumo, setores_exibicao

    diario = (
        periodo.groupby(["Setor", "Data Programada Dia"], as_index=False)[colunas_soma]
        .sum()
    )
    diario["% Atraso"] = 0.0
    mask_diario = diario["Qtde Programada Num"] > 0
    diario.loc[mask_diario, "% Atraso"] = (
        diario.loc[mask_diario, "Realizada Após Num"]
        / diario.loc[mask_diario, "Qtde Programada Num"]
        * 100
    )

    resumo_agrupado = periodo.groupby("Setor")[colunas_soma].sum()
    resumo = pd.DataFrame(index=setores_base).join(resumo_agrupado).fillna(0)
    resumo.index.name = "Setor"
    resumo = resumo.reset_index().rename(
        columns={
            "Qtde Programada Num": "Programada",
            "Realizada no Prazo Num": "Realizada no Prazo",
            "Realizada Após Num": "Realizada Após",
            "Pendente Num": "Pendente",
        }
    )
    resumo["% Atraso Semanal"] = 0.0
    mask_semana = resumo["Programada"] > 0
    resumo.loc[mask_semana, "% Atraso Semanal"] = (
        resumo.loc[mask_semana, "Realizada Após"]
        / resumo.loc[mask_semana, "Programada"]
        * 100
    )

    if somente_movimento:
        resumo = resumo[resumo["Programada"] > 0].copy()

    setores_exibicao = resumo["Setor"].tolist()
    return diario, resumo, setores_exibicao


def classe_celula_calendario(
    programada: float,
    realizada_apos: float,
    pendente: float,
) -> str:
    if programada <= 0:
        return "cal-sem-programacao"
    if realizada_apos > 0:
        return "cal-atraso"
    if pendente > 0:
        return "cal-pendente"
    return "cal-ok"


def html_metricas_calendario(
    programada: float,
    no_prazo: float,
    apos: float,
    pendente: float,
    percentual_atraso: float,
    semanal: bool = False,
) -> str:
    classe = classe_celula_calendario(programada, apos, pendente)
    if programada <= 0:
        return f'<div class="cal-celula {classe}"><span class="cal-vazio">—</span></div>'

    titulo_percentual = "% atraso semanal" if semanal else "% atraso"
    return (
        f'<div class="cal-celula {classe}">'
        f'<div class="cal-linha"><span>Programada</span><strong>{fmt_qtd(programada)}</strong></div>'
        f'<div class="cal-linha"><span>No prazo</span><strong>{fmt_qtd(no_prazo)}</strong></div>'
        f'<div class="cal-linha"><span>Após</span><strong>{fmt_qtd(apos)}</strong></div>'
        f'<div class="cal-linha"><span>Pendente</span><strong>{fmt_qtd(pendente)}</strong></div>'
        f'<div class="cal-linha cal-percentual"><span>{titulo_percentual}</span>'
        f'<strong>{fmt_num(percentual_atraso, 1)}%</strong></div>'
        "</div>"
    )


def renderizar_tabela_calendario_semanal(
    diario: pd.DataFrame,
    resumo: pd.DataFrame,
    setores: list[str],
    inicio: date,
) -> None:
    """Renderiza a matriz setor x dia com rolagem horizontal e cabeçalhos fixos."""
    datas_semana = [inicio + timedelta(days=i) for i in range(5)]

    lookup_diario: dict[tuple[str, date], dict[str, float]] = {}
    for _, linha in diario.iterrows():
        lookup_diario[(str(linha["Setor"]), linha["Data Programada Dia"])] = {
            "programada": float(linha["Qtde Programada Num"]),
            "no_prazo": float(linha["Realizada no Prazo Num"]),
            "apos": float(linha["Realizada Após Num"]),
            "pendente": float(linha["Pendente Num"]),
            "percentual": float(linha["% Atraso"]),
        }

    lookup_resumo = resumo.set_index("Setor").to_dict(orient="index") if not resumo.empty else {}

    cabecalhos = ['<th class="cal-setor-col">Setor</th>']
    for indice, dia in enumerate(datas_semana):
        cabecalhos.append(
            '<th>'
            f'<div class="cal-dia">{DIAS_SEMANA_PT[indice]}</div>'
            f'<div class="cal-data">{dia:%d/%m/%Y}</div>'
            '</th>'
        )
    cabecalhos.append('<th class="cal-resumo-col">Resumo semanal</th>')

    linhas_html: list[str] = []
    for setor in setores:
        celulas = [f'<td class="cal-setor-col"><b>{escape(setor)}</b></td>']
        for dia in datas_semana:
            valores = lookup_diario.get(
                (setor, dia),
                {
                    "programada": 0.0,
                    "no_prazo": 0.0,
                    "apos": 0.0,
                    "pendente": 0.0,
                    "percentual": 0.0,
                },
            )
            celulas.append(
                "<td>"
                + html_metricas_calendario(
                    valores["programada"],
                    valores["no_prazo"],
                    valores["apos"],
                    valores["pendente"],
                    valores["percentual"],
                )
                + "</td>"
            )

        semana = lookup_resumo.get(
            setor,
            {
                "Programada": 0.0,
                "Realizada no Prazo": 0.0,
                "Realizada Após": 0.0,
                "Pendente": 0.0,
                "% Atraso Semanal": 0.0,
            },
        )
        celulas.append(
            '<td class="cal-resumo-col">'
            + html_metricas_calendario(
                float(semana["Programada"]),
                float(semana["Realizada no Prazo"]),
                float(semana["Realizada Após"]),
                float(semana["Pendente"]),
                float(semana["% Atraso Semanal"]),
                semanal=True,
            )
            + "</td>"
        )
        linhas_html.append("<tr>" + "".join(celulas) + "</tr>")

    tabela_html = f"""
    <style>
        .cal-wrapper {{
            overflow-x: auto;
            width: 100%;
            border: 1px solid #cbd5e1;
            border-radius: 10px;
            margin-top: 0.5rem;
            background: #ffffff;
            color-scheme: light;
        }}
        .cal-table {{
            width: max-content;
            min-width: 100%;
            border-collapse: separate;
            border-spacing: 0;
            font-size: 0.88rem;
            color: #111827 !important;
            background: #ffffff !important;
        }}
        .cal-table th, .cal-table td {{
            min-width: 160px;
            padding: 9px;
            border-right: 1px solid #d7dee8;
            border-bottom: 1px solid #d7dee8;
            vertical-align: top;
            background: #ffffff !important;
            color: #111827 !important;
        }}
        .cal-table th {{
            position: sticky;
            top: 0;
            z-index: 3;
            text-align: center;
            background: #e8edf4 !important;
            color: #0f172a !important;
            font-weight: 750;
        }}
        .cal-table .cal-setor-col {{
            min-width: 190px;
            max-width: 260px;
            position: sticky;
            left: 0;
            z-index: 2;
            background: #f1f5f9 !important;
            color: #0f172a !important;
            font-weight: 750;
        }}
        .cal-table th.cal-setor-col {{ z-index: 4; }}
        .cal-table .cal-resumo-col {{
            min-width: 215px;
            background: #eef4ff !important;
            color: #0f172a !important;
        }}
        .cal-dia {{
            font-weight: 800;
            color: #0f172a !important;
        }}
        .cal-data {{
            font-size: 0.79rem;
            color: #475569 !important;
            margin-top: 2px;
            opacity: 1;
        }}
        .cal-celula {{
            border-left: 6px solid transparent;
            border-radius: 8px;
            padding: 8px 9px;
            line-height: 1.35;
            min-height: 128px;
            color: #172033 !important;
            box-shadow: inset 0 0 0 1px rgba(15, 23, 42, 0.04);
        }}
        .cal-celula *, .cal-celula span, .cal-celula strong {{
            color: #172033 !important;
            opacity: 1 !important;
            text-shadow: none !important;
        }}
        .cal-linha {{
            display: flex;
            align-items: baseline;
            justify-content: space-between;
            gap: 10px;
            padding: 2px 0;
            white-space: nowrap;
        }}
        .cal-linha span {{
            font-weight: 650;
        }}
        .cal-linha strong {{
            font-weight: 800;
            text-align: right;
        }}
        .cal-ok {{
            border-left-color: #15803d;
            background: #eaf8ef !important;
        }}
        .cal-pendente {{
            border-left-color: #c96a00;
            background: #fff3e4 !important;
        }}
        .cal-atraso {{
            border-left-color: #d11f2f;
            background: #fdebec !important;
        }}
        .cal-sem-programacao {{
            border-left-color: #8293aa;
            background: #f1f4f8 !important;
            display: flex;
            align-items: center;
            justify-content: center;
        }}
        .cal-vazio {{
            font-size: 1.15rem;
            color: #64748b !important;
            opacity: 1 !important;
        }}
        .cal-percentual {{
            margin-top: 6px;
            padding-top: 6px;
            border-top: 1px solid rgba(15, 23, 42, 0.18);
        }}
        .cal-percentual span, .cal-percentual strong {{
            font-weight: 850;
        }}
        @media (max-width: 900px) {{
            .cal-table {{ font-size: 0.84rem; }}
            .cal-table th, .cal-table td {{ min-width: 155px; }}
            .cal-table .cal-setor-col {{ min-width: 170px; }}
            .cal-celula {{ min-height: 122px; }}
        }}
    </style>
    <div class="cal-wrapper">
        <table class="cal-table">
            <thead><tr>{''.join(cabecalhos)}</tr></thead>
            <tbody>{''.join(linhas_html)}</tbody>
        </table>
    </div>
    """
    st.markdown(tabela_html, unsafe_allow_html=True)


def renderizar_calendario_semanal(df: pd.DataFrame, caminho: Path) -> None:
    st.title("Calendário Semanal dos Setores")
    st.caption(
        "Tela pública organizada pela Data Programada. Cada célula mostra o resultado "
        "do setor naquele dia e a última coluna consolida a semana útil, de segunda a sexta-feira."
    )

    base = preparar_base_calendario_semanal(df)
    limites = limites_datas_programadas(base)
    if limites is None:
        st.warning("Não existem Datas Programadas válidas para montar o calendário.")
        return

    data_min, data_max = limites
    hoje = date.today()
    data_padrao = hoje if data_min <= hoje <= data_max else data_max

    col_data, col_setores, col_movimento = st.columns([1, 2.2, 1.25])
    referencia = col_data.date_input(
        "Escolha uma data da semana",
        value=data_padrao,
        format="DD/MM/YYYY",
        key="calendario_semana_referencia",
        help="O sistema exibirá de segunda a sexta-feira da semana escolhida.",
    )
    setores_disponiveis = opcoes_unicas(base, "Setor")
    setores_sel = col_setores.multiselect(
        "Setores",
        setores_disponiveis,
        placeholder="Vazio = todos os setores",
        key="calendario_setores",
    )
    somente_movimento = col_movimento.checkbox(
        "Somente setores programados",
        value=True,
        key="calendario_somente_movimento",
    )

    inicio = inicio_da_semana(referencia)
    fim = inicio + timedelta(days=4)
    st.info(f"Semana útil exibida: **{inicio:%d/%m/%Y} a {fim:%d/%m/%Y}** (segunda a sexta-feira).")

    diario, resumo, setores_exibicao = montar_dados_calendario_semanal(
        base=base,
        setores_sel=setores_sel,
        inicio=inicio,
        fim=fim,
        somente_movimento=somente_movimento,
    )

    total_programado = float(resumo["Programada"].sum()) if not resumo.empty else 0.0
    total_no_prazo = float(resumo["Realizada no Prazo"].sum()) if not resumo.empty else 0.0
    total_apos = float(resumo["Realizada Após"].sum()) if not resumo.empty else 0.0
    total_pendente = float(resumo["Pendente"].sum()) if not resumo.empty else 0.0
    percentual_atraso = (
        total_apos / total_programado * 100 if total_programado > 0 else 0.0
    )

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Programada na semana", fmt_qtd(total_programado))
    k2.metric("Realizada no prazo", fmt_qtd(total_no_prazo))
    k3.metric("Realizada após", fmt_qtd(total_apos))
    k4.metric("Pendente", fmt_qtd(total_pendente))
    k5.metric("% atraso semanal", f"{fmt_num(percentual_atraso, 1)}%")

    st.caption(
        "Fórmula do atraso semanal: quantidade finalizada após a Data Programada "
        "÷ quantidade programada de segunda a sexta-feira."
    )

    if not setores_exibicao:
        st.warning("Nenhum setor possui programação na semana e nos filtros selecionados.")
        return

    aba_calendario, aba_resumo, aba_detalhes = st.tabs(
        ["Calendário semanal", "Resumo por setor", "Detalhamento"]
    )

    with aba_calendario:
        st.markdown(
            "**Legenda:** 🟢 concluído sem atraso &nbsp;&nbsp; "
            "🟠 possui pendência &nbsp;&nbsp; 🔴 possui realização após o prazo",
            unsafe_allow_html=True,
        )
        renderizar_tabela_calendario_semanal(diario, resumo, setores_exibicao, inicio)

    with aba_resumo:
        resumo_exibicao = resumo.copy()
        resumo_exibicao["Situação"] = "Concluído no prazo"
        resumo_exibicao.loc[resumo_exibicao["Pendente"] > 0, "Situação"] = "Com pendência"
        resumo_exibicao.loc[resumo_exibicao["Realizada Após"] > 0, "Situação"] = "Com atraso"
        resumo_exibicao.loc[resumo_exibicao["Programada"] <= 0, "Situação"] = "Sem programação"

        st.dataframe(
            resumo_exibicao,
            use_container_width=True,
            hide_index=True,
            height=520,
            column_config={
                "Programada": st.column_config.NumberColumn("Programada", format="%.2f"),
                "Realizada no Prazo": st.column_config.NumberColumn(
                    "Realizada no Prazo", format="%.2f"
                ),
                "Realizada Após": st.column_config.NumberColumn(
                    "Realizada Após", format="%.2f"
                ),
                "Pendente": st.column_config.NumberColumn("Pendente", format="%.2f"),
                "% Atraso Semanal": st.column_config.ProgressColumn(
                    "% Atraso Semanal",
                    min_value=0.0,
                    max_value=100.0,
                    format="%.1f%%",
                ),
            },
        )
        csv_resumo = resumo_exibicao.to_csv(
            index=False, sep=";", decimal=",", encoding="utf-8-sig"
        )
        st.download_button(
            "Baixar resumo semanal em CSV",
            data=csv_resumo,
            file_name=f"calendario_semanal_setores_{inicio:%Y%m%d}_{fim:%Y%m%d}.csv",
            mime="text/csv",
        )

    with aba_detalhes:
        detalhes = diario.copy().rename(
            columns={
                "Data Programada Dia": "Data Programada",
                "Qtde Programada Num": "Programada",
                "Realizada no Prazo Num": "Realizada no Prazo",
                "Realizada Após Num": "Realizada Após",
                "Pendente Num": "Pendente",
            }
        )
        if detalhes.empty:
            st.info("Não existem linhas para detalhar nesta semana.")
        else:
            detalhes["Data Programada"] = pd.to_datetime(
                detalhes["Data Programada"]
            ).dt.strftime("%d/%m/%Y")
            detalhes = detalhes[
                [
                    "Data Programada",
                    "Setor",
                    "Programada",
                    "Realizada no Prazo",
                    "Realizada Após",
                    "Pendente",
                    "% Atraso",
                ]
            ]
            st.dataframe(
                detalhes,
                use_container_width=True,
                hide_index=True,
                height=520,
                column_config={
                    "% Atraso": st.column_config.ProgressColumn(
                        "% Atraso",
                        min_value=0.0,
                        max_value=100.0,
                        format="%.1f%%",
                    )
                },
            )

    st.divider()
    st.caption(f"Fonte dos dados: {caminho.name}")


# -----------------------------------------------------------------------------
# Controle de acesso ao apontamento
# -----------------------------------------------------------------------------
def apontamento_autenticado() -> bool:
    """Retorna True somente quando o usuário efetuou login nesta sessão."""
    return bool(st.session_state.get("apontamento_autenticado", False))


def autenticar_apontamento(usuario: str, senha: str) -> bool:
    """Compara as credenciais sem expor o resultado de comparações parciais."""
    usuario_ok = hmac.compare_digest(usuario.strip(), LOGIN_APONTAMENTO)
    senha_ok = hmac.compare_digest(senha, SENHA_APONTAMENTO)
    return usuario_ok and senha_ok


def renderizar_login_apontamento() -> None:
    """Exibe o formulário de login antes de permitir qualquer gravação."""
    st.title("Acesso ao Apontamento de Produção")
    st.caption(
        "As telas Produção por Setor e Calendário Semanal são públicas. "
        "Para lançar ou alterar apontamentos, informe o usuário e a senha."
    )

    col_esquerda, col_login, col_direita = st.columns([1, 1.15, 1])
    with col_login:
        with st.form("form_login_apontamento", clear_on_submit=False):
            usuario = st.text_input(
                "Login",
                placeholder="Digite o login",
                autocomplete="username",
            )
            senha = st.text_input(
                "Senha",
                type="password",
                placeholder="Digite a senha",
                autocomplete="current-password",
            )
            entrar = st.form_submit_button(
                "Entrar no modo apontamento",
                type="primary",
                use_container_width=True,
            )

        if entrar:
            if autenticar_apontamento(usuario, senha):
                st.session_state["apontamento_autenticado"] = True
                st.session_state.pop("erro_login_apontamento", None)
                st.rerun()
            else:
                st.session_state["erro_login_apontamento"] = True

        if st.session_state.get("erro_login_apontamento", False):
            st.error("Login ou senha incorretos.")

        st.info("O acesso permanece ativo somente nesta sessão do navegador.")


def renderizar_status_acesso() -> None:
    """Mostra o estado do acesso e oferece encerramento da sessão protegida."""
    with st.sidebar:
        st.divider()
        st.markdown("### Acesso")
        if apontamento_autenticado():
            st.success("Modo apontamento liberado")
            if st.button("Sair do apontamento", use_container_width=True):
                st.session_state["apontamento_autenticado"] = False
                st.session_state.pop("erro_login_apontamento", None)
                st.rerun()
        else:
            st.info("Modo leitura público")
            st.caption("O apontamento exige login e senha.")


def localizar_arquivo_padrao() -> Path:
    """Localiza a planilha padrão no diretório atual ou ao lado do arquivo Python."""
    candidatos = [
        Path(ARQUIVO_PADRAO).expanduser(),
        Path(__file__).resolve().parent / ARQUIVO_PADRAO,
    ]
    for candidato in candidatos:
        caminho = candidato.resolve()
        if caminho.exists():
            return caminho
    return candidatos[0].resolve()


def carregar_arquivo_modo_leitura() -> tuple[Path, str] | None:
    """No acesso público, usa somente a planilha padrão e não permite trocar arquivos."""
    caminho = localizar_arquivo_padrao()
    with st.sidebar:
        st.header("Fonte dos dados")
        st.caption(f"Arquivo: {caminho.name}")
        st.caption(f"Aba: {NOME_ABA_PADRAO}")

    if not caminho.exists():
        st.error(
            f'Não encontrei o arquivo "{ARQUIVO_PADRAO}". '
            "Coloque a planilha na mesma pasta do aplicativo para liberar a consulta pública."
        )
        return None
    return caminho, NOME_ABA_PADRAO


# -----------------------------------------------------------------------------
# Inicialização do aplicativo
# -----------------------------------------------------------------------------
def rodando_via_streamlit() -> bool:
    """Retorna True quando o arquivo foi iniciado pelo comando streamlit run."""
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx

        return get_script_run_ctx() is not None
    except Exception:
        return False


def abrir_com_streamlit_se_necessario() -> bool:
    """Permite abrir o arquivo pelo PyCharm/duplo clique sem gerar erro de ScriptRunContext."""
    if rodando_via_streamlit():
        return False

    script = Path(__file__).resolve()
    print("\nEste é um aplicativo Streamlit.")
    print("Abrindo corretamente no navegador com o comando:")
    print(f'  "{sys.executable}" -m streamlit run "{script}"\n')

    try:
        subprocess.run([sys.executable, "-m", "streamlit", "run", str(script)], check=False)
    except FileNotFoundError:
        print("Python não encontrado para iniciar o Streamlit.")
    return True


def carregar_arquivo_interface() -> tuple[Path, str] | None:
    with st.sidebar:
        st.header("Arquivo")
        caminho_txt = st.text_input("Caminho do Excel", value=ARQUIVO_PADRAO)
        nome_aba = st.text_input("Aba", value=NOME_ABA_PADRAO)
        st.info("Deixe a planilha Excel fechada antes de salvar, senão o Windows pode bloquear o arquivo.")

    caminho = Path(caminho_txt).expanduser().resolve()

    if not caminho.exists():
        st.warning("Arquivo não encontrado pelo caminho informado.")
        arquivo = st.file_uploader("Ou envie a planilha aqui", type=["xlsx"])
        if arquivo is None:
            st.info("Informe um caminho válido para a planilha ou envie o arquivo Excel acima.")
            st.stop()
            return None
        pasta_tmp = Path("_arquivo_trabalho")
        pasta_tmp.mkdir(exist_ok=True)
        caminho = pasta_tmp / arquivo.name
        conteudo = arquivo.getvalue()
        if not caminho.exists() or caminho.stat().st_size != len(conteudo):
            caminho.write_bytes(conteudo)
            carregar_dados_cacheado.clear()
        st.success(f"Arquivo carregado temporariamente: {caminho}")

    return caminho, nome_aba


def main() -> None:
    st.set_page_config(page_title="Acompanhamento PPCP", layout="wide")

    st.sidebar.success(f"VERSÃO {VERSAO_APP}")
    st.sidebar.markdown("## Navegação")

    pagina_param = str(st.query_params.get("pagina", "")).lower()
    if pagina_param in {"apontamento", "lancamento", "editar"}:
        pagina_padrao = 2
    elif pagina_param in {"calendario", "semanal", "semana"}:
        pagina_padrao = 1
    else:
        pagina_padrao = 0

    pagina = st.sidebar.radio(
        "Escolha a tela",
        [
            "📊 Produção por Setor — leitura",
            "🗓️ Calendário Semanal dos Setores — leitura",
            "🔐 Apontamento de Produção",
        ],
        index=pagina_padrao,
    )
    st.sidebar.caption(
        "As duas telas de consulta são públicas. "
        "O lançamento de apontamentos é protegido pelo login pcp/pcp."
    )
    renderizar_status_acesso()

    # As telas públicas sempre usam a planilha padrão do servidor.
    if "Apontamento" not in pagina:
        arquivo_config = carregar_arquivo_modo_leitura()
        if arquivo_config is None:
            return
        caminho, nome_aba = arquivo_config
    else:
        # Nenhum arquivo é carregado e nenhuma função de gravação é exibida
        # enquanto o login não tiver sido validado.
        if not apontamento_autenticado():
            renderizar_login_apontamento()
            return
        arquivo_config = carregar_arquivo_interface()
        if arquivo_config is None:
            return
        caminho, nome_aba = arquivo_config

    try:
        with st.spinner("Carregando dados da planilha..."):
            df, _ = obter_dados_cacheados(caminho, nome_aba)
    except Exception as exc:
        st.error(f"Erro ao carregar a planilha: {exc}")
        st.stop()
        return

    if df.empty:
        st.warning("Não encontrei linhas de acompanhamento para exibir.")
        st.stop()
        return

    if "Apontamento" in pagina:
        renderizar_apontamento(df, caminho, nome_aba)
    elif "Calendário Semanal" in pagina:
        renderizar_calendario_semanal(df, caminho)
    else:
        renderizar_producao_setor(df, caminho)


if __name__ == "__main__":
    if abrir_com_streamlit_se_necessario():
        raise SystemExit(0)
    main()


## python -m streamlit run app_acompanhamento_ppcp.py

## python -m streamlit run app_acompanhamento_ppcp.py --server.address 172.24.192.1 --server.port 8501
